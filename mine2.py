import logging
import datetime
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Константы состояний для ConversationHandler
REGISTER, DATE, INTERVAL, VIEW, CALC_HOURS, TRANSFER_ACTIVITY, INTERVAL_INPUT, SICK_LEAVE_OPEN_DATE, COLOR_ROWS, NOTIFY_MESSAGE,BAN_USER, BAN_DATE, BAN_REASON, ACTIVITY, RESULT, ENTER_ACTIVITY_USER_ID, ENTER_ACTIVITY_FIO, ENTER_ACTIVITY_DATE, ENTER_ACTIVITY_INTERVAL = range(19)
# Новые константы для редактирования
EDIT_SELECT_ACTIVITY, EDIT_FIELD, EDIT_VALUE, RENAME = range(4)
# Новые константы для удаления
DELETE_SELECT_ACTIVITY, DELETE_CONFIRM, ADD_OLDER_USER, REMOVE_OLDER_USER, DELETE_ANY_ACTIVITY_ROW = range(5)
# Новые константы для ввода данных и удаления
ENTER_DATA_ID, ENTER_DATA_FIO, DELETE_USER_ROW = range(3)
# Путь к Excel-файлу
EXCEL_FILE = 'data.xlsx'

# Список специальных user IDs для доступа к новой функции
SPECIAL_USER_IDS = [461549398,402468895,1352307342]
TARGET_USER_ID = 461549398
DAYS_RU = {
    0: 'понедельник',
    1: 'вторник',
    2: 'среду',
    3: 'четверг',
    4: 'пятницу',
    5: 'субботу',
    6: 'воскресенье',
}
#Инициализирует Excel-файл с двумя листами, если он не существует
def init_excel():
    """Инициализирует Excel-файл с двумя листами, если он не существует."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()

        # Создание первого листа "Users"
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["UserID", "ФИО"])
        # Создание второго листа "Activities"
        ws_activities = wb.create_sheet(title="Activities")
        ws_activities.append(["UserID", "ФИО", "Дата", "Интервал"])
        # Создание листа "Transfers" для переноса активности
        ws_transfers = wb.create_sheet(title="Transfers")
        ws_transfers.append(["UserID", "ФИО", "Дата", "Интервал"])
        #Создание третьего листа "Banned"
        ws_banned = wb.create_sheet("Banned")
        ws_banned.append(["UserID", "ФИО", "Дата окончания бана", "Причина"])
        # Создание четвертого листа "Пройденные активности"
        ws_training = wb.create_sheet("Training")
        ws_training.append(["UserID", "ФИО", "Активность", "Статус"])
        #Пятый лист со старшими
        ws_older_users = wb.create_sheet("OlderUsers")
        ws_older_users.append(["UserID"])

        wb.save(EXCEL_FILE)
#Считывает список зарегистрированных пользователей из листа "Users" Excel-файла
def get_users():
    """
    Считывает список зарегистрированных пользователей из листа "Users" Excel-файла.
    Возвращает список UserID (int).
    """
    if not os.path.exists(EXCEL_FILE):
        init_excel()

    wb = load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]

    users = []
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        user_id = row[0]
        if isinstance(user_id, int):
            users.append(user_id)
    return users

async def send_excel_file(context: ContextTypes.DEFAULT_TYPE):
    try:
        if not os.path.exists(EXCEL_FILE):
            await context.bot.send_message(chat_id=TARGET_USER_ID, text="Файл не найден.")
            return

        with open(EXCEL_FILE, 'rb') as file:
            await context.bot.send_document(chat_id=TARGET_USER_ID, document=file, filename=EXCEL_FILE)
            await context.bot.send_message(chat_id=TARGET_USER_ID, text="Таблица отправлена.")
    except Exception as e:
        logger.error(f"Ошибка при отправке файла: {e}")
        await context.bot.send_message(chat_id=TARGET_USER_ID, text="Произошла ошибка при отправке файла.")

#Обработчик команды /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start. Инициирует регистрацию или показывает главное меню."""
    user_id = update.effective_user.id

    # Инициализируем Excel, если нужно
    init_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]

    # Проверим, зарегистрирован ли пользователь (по UserID)
    registered = False
    fio = None
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            registered = True
            fio = row[1]
            break

    if not registered:
        await update.message.reply_text(
            "Привет! Для начала регистрации, пожалуйста, введи своё ФИО:",
            reply_markup=ReplyKeyboardRemove()
        )
        return REGISTER
    else:
        await update.message.reply_text(
            f"С возвращением, {fio}!",
            reply_markup=main_menu(user_id)
        )
        return ConversationHandler.END
#Создаёт главное меню с кнопками.
def main_menu(user_id):
    """Создаёт главное меню с кнопками. Добавляет 'Скачать таблицу' для специальных пользователей."""
    keyboard = [
        ['Активности', 'Больничный'],
        ['Рассчитать оплату за час', 'Переименовать меня'],
        ['Отметить прохождение теста или обучения']
    ]

    # Добавляем кнопку "Особых действий" только для специальных пользователей
    if user_id in SPECIAL_USER_IDS:
        keyboard.append(['Особые действия'])
    older_users = get_older_users()
    if user_id in older_users:
        keyboard.append(['Меню для старших'])

    # Добавляем кнопку отмены
    keyboard.append(['Отмена'])

    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
#Возвращает пользователя в главное меню
async def back_to_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Возвращает пользователя в главное меню."""
    user_id = update.effective_user.id
    await update.message.reply_text(
        "Главное меню:",
        reply_markup=main_menu(user_id)
    )
#Меню для старших
def senior_menu():
    """Создаёт меню для старших пользователей."""
    keyboard = [
        ['Скачать таблицу', 'Закрасить строки'],
        ['Назад']
    ]
    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
async def senior_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие кнопки 'Меню для старших'."""
    await update.message.reply_text(
        "Меню для старших:",
        reply_markup=senior_menu()
    )
#Создаёт подменю для 'Активности
def activities_menu():
    """Создаёт подменю для 'Активности'."""
    keyboard = [
        ['Добавить активность', 'Перенос активности'],
        ['Внесенные активности', 'Проставленные активности'],
        ['Редактировать активность', 'Удалить активность'],
        ['Назад']
    ]

    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
#Создаёт подменю для 'Больничный'
def sick_menu():
    """Создаёт подменю для 'Больничный'."""
    keyboard = [
        ['Я открыл больничный', 'Я закрыл больничный', 'Больничный продлен'],
        ['Назад']
    ]

    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
#Создаёт подменю для 'Специальных пользователей
def special_menu():
    """Создаёт подменю для 'Специальных пользователей'."""
    keyboard = [
        ['Скачать таблицу', 'Очистить таблицу', 'Внести данные'],
        ['Закрасить строки', 'Оповестить всех', 'Внести в бан'],
        ['Удалить пользователя','Добавить старшего','Удалить старшего'],
        ['Удалить активити','Внести активность'],
        ['Назад']
    ]
    return ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
#Старшие спецы сохранены тут
def get_older_users():
    """Получает список ID пользователей OLDER_USERS из листа 'OlderUsers'."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_older_users = wb["OlderUsers"]
    older_users = []
    for row in ws_older_users.iter_rows(min_row=2, values_only=True):
        user_id = row[0]
        if isinstance(user_id, int):
            older_users.append(user_id)
    return older_users
def add_older_user(user_id):
    """Добавляет ID пользователя в лист 'OlderUsers'."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_older_users = wb["OlderUsers"]
    ws_older_users.append([user_id])
    wb.save(EXCEL_FILE)
def remove_older_user(user_id):
    """Удаляет ID пользователя из листа 'OlderUsers'."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_older_users = wb["OlderUsers"]
    for row in ws_older_users.iter_rows(min_row=2, max_row=ws_older_users.max_row):
        if row[0].value == user_id:
            ws_older_users.delete_rows(row[0].row, 1)
            wb.save(EXCEL_FILE)
            break
# Обработчик для начала процесса ввода данных пользователя на лист "Activities"
async def enter_activity_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс ввода данных пользователя на лист 'Activities'."""
    await update.message.reply_text(
        "Введите ID пользователя:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ENTER_ACTIVITY_USER_ID
# Функция для ввода ID пользователя
async def enter_activity_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет ID пользователя и запрашивает ФИО."""
    user_id = update.message.text.strip()
    if not user_id.isdigit():
        await update.message.reply_text(
            "ID должен быть числом. Пожалуйста, введите корректный ID:"
        )
        return ENTER_ACTIVITY_USER_ID
    context.user_data['enter_activity_user_id'] = int(user_id)
    await update.message.reply_text(
        "Введите ФИО пользователя:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ENTER_ACTIVITY_FIO
# Функция для ввода ФИО пользователя
async def enter_activity_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет ФИО пользователя и запрашивает дату."""
    fio = update.message.text.strip()
    if not fio:
        await update.message.reply_text(
            "ФИО не может быть пустым. Пожалуйста, введите ФИО пользователя:"
        )
        return ENTER_ACTIVITY_FIO
    context.user_data['enter_activity_fio'] = fio
    await update.message.reply_text(
        "Введите дату активности в формате ДД.ММ.ГГГГ:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ENTER_ACTIVITY_DATE
# Функция для ввода даты активности
async def enter_activity_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет дату и запрашивает интервал."""
    date_text = update.message.text.strip()
    if not validate_date(date_text):
        await update.message.reply_text(
            "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ:"
        )
        return ENTER_ACTIVITY_DATE
    context.user_data['enter_activity_date'] = date_text
    await update.message.reply_text(
        "Пожалуйста, введите интервал активности (пример: СВУ с 12 до 13:45):"
    )
    return ENTER_ACTIVITY_INTERVAL
# Функция для ввода интервала активности
async def enter_activity_interval(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет интервал и записывает данные в Excel."""
    interval = update.message.text.strip()
    if not interval:
        await update.message.reply_text(
            "Интервал не может быть пустым. Пожалуйста, введите интервал активности:"
        )
        return ENTER_ACTIVITY_INTERVAL
    user_id = context.user_data.get('enter_activity_user_id')
    fio = context.user_data.get('enter_activity_fio')
    date = context.user_data.get('enter_activity_date')
    if not all([user_id, fio, date]):
        await update.message.reply_text(
            "Ошибка при получении данных. Попробуйте снова.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
    # Записываем данные в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]
    ws_activities.append([user_id, fio, date, interval])
    wb.save(EXCEL_FILE)
    await update.message.reply_text(
        f"Данные пользователя с ID {user_id} успешно внесены!",
        reply_markup=main_menu(update.effective_user.id)
    )
    return ConversationHandler.END
async def add_older_user_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс добавления ID пользователя в OLDER_USERS."""
    await update.message.reply_text(
        "Введите ID пользователя для добавления в OLDER_USERS:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ADD_OLDER_USER
async def add_older_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавляет ID пользователя в OLDER_USERS."""
    user_id = update.message.text.strip()
    if not user_id.isdigit():
        await update.message.reply_text(
            "ID должен быть числом. Пожалуйста, введите корректный ID:"
        )
        return ADD_OLDER_USER
    user_id = int(user_id)
    add_older_user(user_id)
    await update.message.reply_text(
        f"ID пользователя {user_id} успешно добавлен в OLDER_USERS.",
        reply_markup=main_menu(update.effective_user.id)
    )
    return ConversationHandler.END
async def remove_older_user_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс удаления ID пользователя из OLDER_USERS."""
    await update.message.reply_text(
        "Введите ID пользователя для удаления из OLDER_USERS:",
        reply_markup=ReplyKeyboardRemove()
    )
    return REMOVE_OLDER_USER
async def remove_older_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет ID пользователя из OLDER_USERS."""
    user_id = update.message.text.strip()
    if not user_id.isdigit():
        await update.message.reply_text(
            "ID должен быть числом. Пожалуйста, введите корректный ID:"
        )
        return REMOVE_OLDER_USER
    user_id = int(user_id)
    remove_older_user(user_id)
    await update.message.reply_text(
        f"ID пользователя {user_id} успешно удален из OLDER_USERS.",
        reply_markup=main_menu(update.effective_user.id)
    )
    return ConversationHandler.END
#Сохраняет ФИО пользователя и завершает регистрацию
async def register_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет ФИО пользователя и завершает регистрацию."""
    fio = update.message.text.strip()
    if not fio:
        await update.message.reply_text(
            "ФИО не может быть пустым. Пожалуйста, введите своё ФИО:"
        )
        return REGISTER

    user_id = update.effective_user.id

    # Записываем ФИО в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]

    # Проверим ещё раз, зарегистрирован ли пользователь
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            await update.message.reply_text(
                "Вы уже зарегистрированы.",
                reply_markup=main_menu(user_id)
            )
            return ConversationHandler.END

    ws_users.append([user_id, fio])

    wb.save(EXCEL_FILE)

    await update.message.reply_text(
        f"Спасибо за регистрацию, {fio}!",
        reply_markup=main_menu(user_id)
    )
    return ConversationHandler.END
#Внесение данных за пользователя
async def enter_data_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс ввода данных пользователя."""
    await update.message.reply_text(
        "Введите ID пользователя:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ENTER_DATA_ID
async def enter_data_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет ID пользователя и запрашивает ФИО."""
    user_id = update.message.text.strip()
    if not user_id.isdigit():
        await update.message.reply_text(
            "ID должен быть числом. Пожалуйста, введите корректный ID:"
        )
        return ENTER_DATA_ID
    context.user_data['enter_user_id'] = int(user_id)
    await update.message.reply_text(
        "Введите ФИО пользователя:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ENTER_DATA_FIO
async def enter_data_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет ФИО пользователя и завершает ввод данных."""
    fio = update.message.text.strip()
    if not fio:
        await update.message.reply_text(
            "ФИО не может быть пустым. Пожалуйста, введите ФИО пользователя:"
        )
        return ENTER_DATA_FIO
    user_id = context.user_data.get('enter_user_id')
    if not user_id:
        await update.message.reply_text(
            "Ошибка при получении ID пользователя. Попробуйте снова.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
    # Записываем данные в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    ws_users.append([user_id, fio])
    wb.save(EXCEL_FILE)
    await update.message.reply_text(
        f"Данные пользователя с ID {user_id} успешно внесены!",
        reply_markup=main_menu(update.effective_user.id)
    )
    return ConversationHandler.END
#Удаление любой активности в файле
async def delete_any_activity_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс удаления строки на листе 'Activities'."""
    await update.message.reply_text(
        "Введите номер строки, которую хотите удалить:",
        reply_markup=ReplyKeyboardRemove()
    )
    return DELETE_ANY_ACTIVITY_ROW
# Функция для удаления строки на листе "Activities"
async def delete_any_activity_row(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет строку на листе 'Activities' по указанному номеру."""
    row_number = update.message.text.strip()
    if not row_number.isdigit():
        await update.message.reply_text(
            "Номер строки должен быть числом. Пожалуйста, введите корректный номер строки:"
        )
        return DELETE_ANY_ACTIVITY_ROW

    row_number = int(row_number)
    if row_number < 2:
        await update.message.reply_text(
            "Минимально допустимый номер строки - 2."
        )
        return DELETE_ANY_ACTIVITY_ROW

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_activities = wb["Activities"]
        max_row = ws_activities.max_row
        if row_number > max_row:
            await update.message.reply_text(
                f"В таблице только {max_row} строк. Введите корректный номер строки."
            )
            return DELETE_ANY_ACTIVITY_ROW

        ws_activities.delete_rows(row_number, 1)
        wb.save(EXCEL_FILE)
        await update.message.reply_text(
            f"Строка {row_number} успешно удалена.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Ошибка при удалении строки: {e}")
        await update.message.reply_text(
            "Произошла ошибка при удалении строки. Попробуйте снова.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
#Удаление пользователя
async def delete_user_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс удаления пользователя."""
    await update.message.reply_text(
        "Введите номер строки пользователя, которого хотите удалить:",
        reply_markup=ReplyKeyboardRemove()
    )
    return DELETE_USER_ROW
async def delete_user_row(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет пользователя по номеру строки."""
    row_number = update.message.text.strip()
    if not row_number.isdigit():
        await update.message.reply_text(
            "Номер строки должен быть числом. Пожалуйста, введите корректный номер строки:"
        )
        return DELETE_USER_ROW
    row_number = int(row_number)
    if row_number < 2:
        await update.message.reply_text(
            "Минимально допустимый номер строки - 2."
        )
        return DELETE_USER_ROW
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_users = wb["Users"]
        max_row = ws_users.max_row
        if row_number > max_row:
            await update.message.reply_text(
                f"В таблице только {max_row} строк. Введите корректный номер строки."
            )
            return DELETE_USER_ROW
        ws_users.delete_rows(row_number, 1)
        wb.save(EXCEL_FILE)
        await update.message.reply_text(
            f"Пользователь в строке {row_number} успешно удален.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"Ошибка при удалении пользователя: {e}")
        await update.message.reply_text(
            "Произошла ошибка при удалении пользователя. Попробуйте снова.",
            reply_markup=main_menu(update.effective_user.id)
        )
        return ConversationHandler.END
#Инициализирует процесс переименования пользователя.
async def rename_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Инициализирует процесс переименования пользователя."""
    user_id = update.effective_user.id

    # Проверяем, зарегистрирован ли пользователь
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    registered = False

    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            registered = True

    if not registered:
        await update.message.reply_text(
            "Вы не зарегистрированы. Сначала пройдите регистрацию с помощью команды /start."
        )
        return ConversationHandler.END

    await update.message.reply_text(
        "Пожалуйста, введите новое ФИО:",
        reply_markup=ReplyKeyboardRemove()  # Убираем клавиатуру
    )
    return RENAME  # Переключаемся на состояние PANE
#Обновляет ФИО пользователя в Excel.
async def update_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обновляет ФИО пользователя в Excel."""
    new_fio = update.message.text.strip()
    if not new_fio:
        await update.message.reply_text(
            "ФИО не может быть пустым. Пожалуйста, введите своё ФИО:"
        )
        return RENAME

    user_id = update.effective_user.id

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]

    # Находим пользователя и меняем его ФИО
    for row in ws_users.iter_rows(min_row=2):
        if row[0].value == user_id:
            row[1].value = new_fio  # Обновляем ФИО
            wb.save(EXCEL_FILE)
            await update.message.reply_text(
                f"Ваше ФИО успешно обновлено на: {new_fio}",
                reply_markup=ReplyKeyboardMarkup([["Назад"]], resize_keyboard=True)  # Восстанавливаем клавиатуру
            )
            return ConversationHandler.END

    await update.message.reply_text(
        "Ошибка при обновлении ФИО. Попробуйте снова.",
        reply_markup=ReplyKeyboardMarkup([["Назад"]], resize_keyboard=True)
    )
    return ConversationHandler.END
#Получает ФИО пользователя по его user_id
def get_user_fio(user_id):
    """Получает ФИО пользователя по его user_id."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb['Users']
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            return row[1]
    return None
#Начинает процесс добавления активности.
async def add_activity_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс добавления активности."""
    user_id = update.effective_user.id
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    ws_banned = wb["Banned"]

    fio = None
    registered = False
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            registered = True
            fio = row[1]
            break

    if not registered:
        await update.message.reply_text(
            "Пожалуйста, зарегистрируйтесь сначала, отправив /start."
        )
        return ConversationHandler.END

    # Проверка бана
    today = datetime.today().date()
    is_banned = False
    ban_end_date = None
    for row in ws_banned.iter_rows(min_row=2, values_only=True):
        banned_user_id, _, end_date_str, _ = row
        if banned_user_id == user_id:
            try:
                end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
                if today <= end_date:
                    is_banned = True
                    ban_end_date = end_date
                    break
            except ValueError:
                continue

    if is_banned:
        await update.message.reply_text(
            f"Вы забанены до {ban_end_date.strftime('%d.%m.%Y')}. "
            "Вы не можете добавлять активности до окончания бана."
        )
        return ConversationHandler.END

    context.user_data['fio'] = fio  # Сохраняем ФИО пользователя для дальнейшего использования

    await update.message.reply_text(
        "Пожалуйста, введите дату активности в формате ДД.ММ.ГГГГ:",
        reply_markup=ReplyKeyboardRemove()
    )
    return DATE
#Сохраняет дату и запрашивает интервал.
async def add_activity_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет дату и запрашивает интервал."""
    date_text = update.message.text.strip()
    if not validate_date(date_text):
        await update.message.reply_text(
            "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ:"
        )
        return DATE

    context.user_data['date'] = date_text
    await update.message.reply_text(
        "Пожалуйста, введите интервал активности (пример: СВУ с 12 до 13:45):"
    )
    return INTERVAL
#Сохраняет интервал и записывает данные в Excel.
async def add_activity_interval(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет интервал и записывает данные в Excel."""
    interval = update.message.text.strip()
    date = context.user_data.get('date')
    fio = context.user_data.get('fio')

    if not interval:
        await update.message.reply_text(
            "Интервал не может быть пустым. Пожалуйста, введите интервал активности:"
        )
        return INTERVAL
    try:
            activity_date = datetime.strptime(date, '%d.%m.%Y')
    except ValueError:
            await update.message.reply_text(
                "Неверный формат даты. Пожалуйста, начните процесс заново."
            )
            return ConversationHandler.END

# Получаем числовое представление дня недели
    weekday_num = activity_date.weekday()  # 0=понедельник, 6=воскресенье

# Проверка, является ли дата субботой, воскресеньем или понедельником
    if weekday_num in (0, 5, 6):  # 0=понедельник, 5=суббота, 6=воскресенье
        day_ru = DAYS_RU.get(weekday_num, 'неизвестный день')

        notification_text = f"Пользователь {fio} добавил активность на {day_ru}, {date}."

        for special_user_id in SPECIAL_USER_IDS:
            try:
                await context.bot.send_message(
                    chat_id=special_user_id,
                    text=notification_text
                )
            except Exception as e:
                print(f"Не удалось отправить сообщение пользователю {special_user_id}: {e}")

    # Записываем данные в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]

    user_id = update.effective_user.id
    ws_activities.append([user_id, fio, date, interval])
    wb.save(EXCEL_FILE)

    await update.message.reply_text(
        "Активность добавлена успешно!",
        reply_markup=main_menu(user_id)
    )

    return ConversationHandler.END
#Запрашивает информацию об активности или обучении
async def activity_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает информацию об активности или обучении."""
    user_id = update.effective_user.id
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]

    registered = False
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            registered = True
            break

    if not registered:
        await update.message.reply_text(
            "Пожалуйста, зарегистрируйтесь сначала, отправив /start."
        )
        return ConversationHandler.END

    await update.message.reply_text(
        "Какую активность или обучение вы прошли/выполнили?",
        reply_markup=ReplyKeyboardRemove()  # Убираем клавиатуру
    )
    return ACTIVITY  # Переключаемся на состояние ACTIVITY
#Обрабатывает введенную активность и запрашивает результат
async def process_activity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенную активность и запрашивает результат."""
    activity = update.message.text.strip()

    if not activity:
        await update.message.reply_text("Активность не может быть пустой. Пожалуйста, попробуйте снова.")
        return ACTIVITY

    user_id = update.effective_user.id
    # Получаем ФИО пользователя из листа "Users"
    fio = None

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            fio = row[1]
            break
    wb.close()

    if not fio:
        await update.message.reply_text("Не удалось найти ваше ФИО в базе данных.")
        return ConversationHandler.END

    # Записываем активность в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_training = wb["Training"]

    # Добавляем новую активность
    ws_training.append([user_id, fio, activity, None])  # Теперь записываем ФИО
    wb.save(EXCEL_FILE)

    await update.message.reply_text(
        "Каков результат?",
        reply_markup=ReplyKeyboardMarkup(
            [["Пройдено"], ["Не пройдено"]],
            resize_keyboard=True
        )
    )
    return RESULT  # Переключаемся на состояние RESULT
#Обрабатывает результат активности и завершает процесс.
async def process_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает результат активности и завершает процесс."""
    result = update.message.text.strip()

    if result not in ["Пройдено", "Не пройдено"]:
        await update.message.reply_text("Пожалуйста, выберите 'Пройдено' или 'Не пройдено'.")
        return RESULT

    user_id = update.effective_user.id

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_training = wb["Training"]

    # Преобразуем генератор в список
    rows = list(ws_training.iter_rows(min_row=2))

    # Находим последнюю запись для данного пользователя и обновляем статус
    for row in reversed(rows):
        if row[0].value == user_id:
            row[3].value = result  # Обновляем статус
            wb.save(EXCEL_FILE)
            break

    await update.message.reply_text(
        f"Вы отметили активность как: {result}.",
        reply_markup=ReplyKeyboardMarkup([["Назад"]], resize_keyboard=True)  # Восстанавливаем клавиатуру
    )

    return ConversationHandler.END
#Показывает проставленные активности
async def show_recorded_activities(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]

    recorded_activities = []

    # Определите, какая колонка отвечает за закрашенные ячейки.
    # Предположим, что это колонка "Interval" (например, 4-я колонка)
    INTERVAL_COLUMN_INDEX = 1

    for row in ws_activities.iter_rows(min_row=2):  # Пропустить заголовок
        row_user_id = row[0].value
        if row_user_id != user_id:
            continue

        interval_cell = row[INTERVAL_COLUMN_INDEX - 1]  # Индексация с 0
        fill = interval_cell.fill

        # Проверим, есть ли цвет заливки (исключим пустую заливку)
        if fill and fill.fgColor and fill.fgColor.type != 'none' and fill.fgColor.rgb != '00000000':
            # Получаем необходимые данные, например: ФИО, дату и интервал
            fio = row[1].value
            date = row[2].value
            interval = row[3].value
            recorded_activities.append(f"Дата: {date}, Время: {interval}")

    if not recorded_activities:
        await update.message.reply_text("Пока ничего не проставили.")
    else:
        activities_text = "\n".join(recorded_activities)
        await update.message.reply_text(f"Проставлено:\n{activities_text}", reply_markup=main_menu(user_id))

    wb.close()
#Начинает процесс закрашивания строк
async def color_rows_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс закрашивания строк."""
    await update.message.reply_text(
        "Введите номер строки, до которой необходимо закрасить данные (целое число):",
        reply_markup=ReplyKeyboardRemove()
    )
    return COLOR_ROWS
#Закрашивает строки до указанного номера включительно в желтый цвет.
async def color_rows_process(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Закрашивает строки до указанного номера включительно в желтый цвет."""
    row_input = update.message.text.strip()

    if not row_input.isdigit():
        await update.message.reply_text(
            "Пожалуйста, введите корректное целое число для номера строки:"
        )
        return COLOR_ROWS

    row_number = int(row_input)

    if row_number < 2:
        await update.message.reply_text(
            "Минимально допустимый номер строки - 2."
        )
        return COLOR_ROWS

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_activities = wb["Activities"]

        max_row = ws_activities.max_row
        if row_number > max_row:
            await update.message.reply_text(
                f"В таблице только {max_row} строк. Будут закрашены все строки до {max_row}."
            )
            row_number = max_row

        # Определяем желтый цвет
        yellow_fill = PatternFill(start_color="FFFF00",
                                  end_color="FFFF00",
                                  fill_type="solid")

        # Закрашиваем строки от 2 до row_number включительно (предполагая, что 1-я строка - заголовок)
        for row in ws_activities.iter_rows(min_row=2, max_row=row_number, max_col=ws_activities.max_column):
            for cell in row:
                cell.fill = yellow_fill

        wb.save(EXCEL_FILE)

        await update.message.reply_text(
            f"Строки с 2 по {row_number} успешно закрашены в желтый цвет.",
            reply_markup=main_menu(update.effective_user.id)
        )

        return ConversationHandler.END

    except Exception as e:
        logger.error(f"Ошибка при закрашивании строк: {e}")
        await update.message.reply_text(
            "Произошла ошибка при обработке запроса. Пожалуйста, попробуйте позже."
        )
        return ConversationHandler.END
#Обрабатывает нажатие кнопок главного меню.
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие кнопок главного меню."""
    text = update.message.text
    user_id = update.effective_user.id

    if text == 'Активности':
        await update.message.reply_text(
            "Выберите действие в разделе 'Активности':",
            reply_markup=activities_menu()
        )
    elif text == 'Больничный':
        await update.message.reply_text(
            "Выберите действие в разделе 'Больничный':",
            reply_markup=sick_menu()
        )
    elif text == 'Особые действия':
        await update.message.reply_text(
            "Выберите действие в разделе для специальных пользователей:",
            reply_markup=special_menu()
        )
    elif text == 'Меню для старших':
        await update.message.reply_text(
            "Выберите действие в разделе для старших специалистов:",
            reply_markup=senior_menu()
        )
    elif text == 'Оповестить всех':
        await update.message.reply_text(
            "Введите текст для отправки всем пользователям:",
            reply_markup=ReplyKeyboardRemove()
        )
        return NOTIFY_MESSAGE
    elif text == 'Скачать таблицу':
        return await download_table(update, context)
    elif text == 'Назад':
        await back_to_main_menu(update, context)
    elif text == 'Отмена':
        await cancel(update, context)
    else:
        await update.message.reply_text(
            "Пожалуйста, используйте кнопки меню.",
            reply_markup=main_menu(user_id)
        )
#Инициирует процесс переноса активности, запрашивая интервал.
async def transfer_activity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Инициирует процесс переноса активности, запрашивая интервал."""
    await update.message.reply_text(
        "Введите интервал переноса (пример: с 7:00 на 7:30):",
        reply_markup=ReplyKeyboardRemove()
    )
    return INTERVAL_INPUT
#БЭКККАААААП
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    file = await update.message.document.get_file()
    upload_dir = "uploads"
    os.makedirs(upload_dir, exist_ok=True)  # Создаем директорию, если она не существует
    file_path = os.path.join(upload_dir, f"{user_id}_{update.message.document.file_name}")
    await file.download_to_drive(file_path)
    await update.message.reply_text("Файл получен. Начинаю обработку...")
    await process_uploaded_file(file_path, update, context)
async def process_uploaded_file(file_path, update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        # Загрузка загруженного файла
        uploaded_wb = load_workbook(file_path)
        # Загрузка инициализированного файла
        init_wb = load_workbook(EXCEL_FILE)

        # Перенос данных из листа "Users"
        if "Users" in uploaded_wb.sheetnames:
            uploaded_ws_users = uploaded_wb["Users"]
            init_ws_users = init_wb["Users"]
            for row in uploaded_ws_users.iter_rows(min_row=2, values_only=True):
                init_ws_users.append(row)

        # Перенос данных из листа "Activities"
        if "Activities" in uploaded_wb.sheetnames:
            uploaded_ws_activities = uploaded_wb["Activities"]
            init_ws_activities = init_wb["Activities"]
            for row in uploaded_ws_activities.iter_rows(min_row=2, values_only=True):
                init_ws_activities.append(row)

        # Перенос данных из листа "Banned"
        if "Banned" in uploaded_wb.sheetnames:
            uploaded_ws_banned = uploaded_wb["Banned"]
            init_ws_banned = init_wb["Banned"]
            for row in uploaded_ws_banned.iter_rows(min_row=2, values_only=True):
                init_ws_banned.append(row)

        # Перенос данных из листа "Training"
        if "Training" in uploaded_wb.sheetnames:
            uploaded_ws_training = uploaded_wb["Training"]
            init_ws_training = init_wb["Training"]
            for row in uploaded_ws_training.iter_rows(min_row=2, values_only=True):
                init_ws_training.append(row)

        # Сохранение изменений
        init_wb.save(EXCEL_FILE)
        await update.message.reply_text("Данные успешно перенесены.")
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}")
        await update.message.reply_text("Произошла ошибка при обработке файла. Пожалуйста, попробуйте позже.")
    finally:
        # Удаление временного файла
        os.remove(file_path)
#Начинает процесс отправки уведомления всем пользователям.
async def notify_all_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс отправки уведомления всем пользователям."""
    user_id = update.effective_user.id
    if user_id not in SPECIAL_USER_IDS:
        await update.message.reply_text("У вас нет прав для выполнения этого действия.")
        return ConversationHandler.END

    await update.message.reply_text(
        "Введите текст для отправки всем пользователям:",
        reply_markup=ReplyKeyboardRemove()
    )
    return NOTIFY_MESSAGE
#Отправляет введённое сообщение всем зарегистрированным пользователям
async def notify_all_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет введённое сообщение всем зарегистрированным пользователям."""
    user_id = update.effective_user.id
    text = update.message.text

    users = get_users()

    if not users:
        await update.message.reply_text("Список пользователей пуст.")
        return ConversationHandler.END

    sent_count = 0
    failed_users = []

    for uid in users:
        try:
            await context.bot.send_message(chat_id=uid, text=text)
            sent_count += 1
        except Exception as e:
            failed_users.append(uid)
            print(f"Не удалось отправить сообщение пользователю {uid}: {e}")

    response = f"Сообщение отправлено {sent_count} из {len(users)} пользователей."
    if failed_users:
        response += f"\nНе удалось отправить сообщение следующим пользователям: {failed_users}"
    await update.message.reply_text(response)

    # Возвращаемся в главное меню
    await back_to_main_menu(update, context)
    return ConversationHandler.END
#Получает введенный интервал и сохраняет его
async def interval_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получает введенный интервал и сохраняет его."""
    interval = update.message.text.strip()
    user_id = update.effective_user.id
    fio = get_user_fio(user_id)
    date = datetime.now()

    if not interval:
        await update.message.reply_text(
            "Интервал не может быть пустым. Пожалуйста, введите интервал переноса (пример: с 7:00 на 7:30):"
        )
        return INTERVAL_INPUT

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_transfers = wb['Transfers']

    ws_transfers.append([user_id, fio, date.strftime('%Y-%m-%d %H:%M:%S'), interval])
    wb.save(EXCEL_FILE)

    await update.message.reply_text("Информация передана ВС.", reply_markup=main_menu(user_id))
    return ConversationHandler.END
#Начинает процесс просмотра активностей
async def view_activities_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс просмотра активностей."""
    user_id = update.effective_user.id

    wb = openpyxl.load_workbook(EXCEL_FILE)

    ws_activities = wb["Activities"]

    user_activities = []
    for row in ws_activities.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_activities.append(row)

    if not user_activities:
        await update.message.reply_text(
            "У вас пока нет добавленных активностей.",
            reply_markup=main_menu(user_id)
        )
        return ConversationHandler.END

    message = "Ваши активности:\n\n"
    for idx, activity in enumerate(user_activities, start=1):
        date = activity[2]
        interval = activity[3]
        message += f"{idx}. Дата: {date}\n   Интервал: {interval}\n\n"

    await update.message.reply_text(message, reply_markup=main_menu(user_id))
    return ConversationHandler.END
#Начало процесса редактирования активности
async def edit_activity_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало процесса редактирования активности."""
    user_id = update.effective_user.id
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]

    # Получаем список активностей пользователя
    user_activities = [

        (idx, row) for idx, row in enumerate(ws_activities.iter_rows(min_row=2, values_only=True), start=2)
        if row[0] == user_id
    ]

    wb.close()

    if not user_activities:
        await update.message.reply_text(
            "У вас нет активностей для редактирования.",
            reply_markup=main_menu(user_id)
        )
        return ConversationHandler.END

    # Формируем сообщение с перечнем активностей
    message = "Выберите номер активности для редактирования:\n"
    for idx, (row_num, activity) in enumerate(user_activities, start=1):
        message += f"{idx}. Дата: {activity[2]}, Интервал: {activity[3]}\n"

    # Сохраняем номера строк для дальнейшего использования
    context.user_data['edit_user_activities'] = user_activities

    await update.message.reply_text(message)
    return EDIT_SELECT_ACTIVITY
#Выбор активности для редактирования
async def edit_select_activity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выбор активности для редактирования."""
    try:
        selection = int(update.message.text)
        user_activities = context.user_data.get('edit_user_activities', [])

        if 1 <= selection <= len(user_activities):
            selected_activity = user_activities[selection - 1]
            context.user_data['selected_activity_row'] = selected_activity[0]
            await update.message.reply_text(
                "Выберите поле для редактирования (Дата/Интервал):",
                reply_markup=ReplyKeyboardMarkup([['Дата', 'Интервал']], one_time_keyboard=True, resize_keyboard=True)
            )
            return EDIT_FIELD
        else:
            await update.message.reply_text("Неверный выбор. Пожалуйста, введите корректный номер активности.")
            return EDIT_SELECT_ACTIVITY

    except ValueError:
        await update.message.reply_text("Пожалуйста, введите числовой номер активности.")
        return EDIT_SELECT_ACTIVITY
#Выбор поля для редактирования
async def edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выбор поля для редактирования."""
    field = update.message.text
    if field not in ['Дата', 'Интервал']:
        await update.message.reply_text("Пожалуйста, выберите либо 'Дата', либо 'Интервал'.")
        return EDIT_FIELD

    context.user_data['edit_field'] = field
    await update.message.reply_text(f"Введите новое значение для {field}:")
    return EDIT_VALUE
#Ввод нового значения
async def edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ввод нового значения для выбранного поля."""
    new_value = update.message.text.strip()
    field = context.user_data.get('edit_field')
    row = context.user_data.get('selected_activity_row')

    if not all([new_value, field, row]):
        await update.message.reply_text("Произошла ошибка при получении данных. Попробуйте снова.",
                                        reply_markup=main_menu(update.effective_user.id))
        return ConversationHandler.END

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]

    if field == 'Дата':
        # Проверка формата даты
        try:
            datetime.strptime(new_value, "%d.%m.%Y")
        except ValueError:
            await update.message.reply_text("Неверный формат даты. Пожалуйста, используйте ДД.ММ.ГГГГ.")
            wb.close()
            return EDIT_VALUE
        ws_activities.cell(row=row, column=3).value = new_value
    elif field == 'Интервал':
        ws_activities.cell(row=row, column=4).value = new_value
    for cell in ws_activities[row]:
        cell.fill = PatternFill(fill_type=None)

    wb.save(EXCEL_FILE)
    wb.close()

    await update.message.reply_text(f"Активность успешно обновлена.\n{field}: {new_value}",
                                    reply_markup=main_menu(update.effective_user.id))
    return ConversationHandler.END
#Выбор активности для удаления
async def delete_activity_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало процесса удаления активности."""
    user_id = update.effective_user.id
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_activities = wb["Activities"]

    # Получаем список активностей пользователя
    user_activities = [

        (idx, row) for idx, row in enumerate(ws_activities.iter_rows(min_row=2, values_only=True), start=2)
        if row[0] == user_id
    ]

    wb.close()

    if not user_activities:
        await update.message.reply_text(
            "У вас нет активностей для удаления.",
            reply_markup=main_menu(user_id)
        )
        return ConversationHandler.END

    # Формируем сообщение с перечнем активностей
    message = "Выберите номер активности для удаления:\n"
    for idx, (row_num, activity) in enumerate(user_activities, start=1):
        message += f"{idx}. Дата: {activity[2]}, Интервал: {activity[3]}\n"

    # Сохраняем номера строк для дальнейшего использования
    context.user_data['delete_user_activities'] = user_activities

    await update.message.reply_text(message, reply_markup=ReplyKeyboardMarkup([['Да', 'Нет']], one_time_keyboard=True,
                                                                              resize_keyboard=True))
    return DELETE_SELECT_ACTIVITY
#Выбор активности для удаления
async def delete_select_activity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выбор активности для удаления."""
    try:
        selection = int(update.message.text)
        user_activities = context.user_data.get('delete_user_activities', [])

        if 1 <= selection <= len(user_activities):
            selected_activity = user_activities[selection - 1]
            context.user_data['selected_delete_row'] = selected_activity[0]
            # Подтверждение удаления
            await update.message.reply_text(
                "Вы уверены, что хотите удалить эту активность? (Да/Нет)",
                reply_markup=ReplyKeyboardMarkup([['Да', 'Нет']], one_time_keyboard=True, resize_keyboard=True)
            )
            return DELETE_CONFIRM
        else:
            await update.message.reply_text("Неверный выбор. Пожалуйста, введите корректный номер активности.")
            return DELETE_SELECT_ACTIVITY

    except ValueError:
        await update.message.reply_text("Пожалуйста, введите числовой номер активности.")
        return DELETE_SELECT_ACTIVITY
#Подтверждение удаления активности
async def delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Подтверждение удаления активности."""
    confirmation = update.message.text.lower()
    user_id = update.effective_user.id

    if confirmation == 'да':
        row = context.user_data.get('selected_delete_row')

        if not row:
            await update.message.reply_text("Произошла ошибка при получении данных. Попробуйте снова.",
                                            reply_markup=main_menu(user_id))
            return ConversationHandler.END

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_activities = wb["Activities"]
        ws_activities.delete_rows(row, 1)
        wb.save(EXCEL_FILE)
        wb.close()

        await update.message.reply_text("Активность успешно удалена.", reply_markup=main_menu(user_id))
        return ConversationHandler.END
    elif confirmation == 'нет':
        await update.message.reply_text("Удаление активности отменено.", reply_markup=main_menu(user_id))
        return ConversationHandler.END
    else:
        await update.message.reply_text("Пожалуйста, ответьте 'Да' или 'Нет'.")
        return DELETE_CONFIRM
#начало бана
async def ban_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Введите ID пользователя для бана:",
        reply_markup=ReplyKeyboardRemove()
    )
    return BAN_USER
#обработка ID для бана
async def ban_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.text.strip()
    if not user_id.isdigit():
        await update.message.reply_text("ID должен быть числом. Попробуйте снова:")
        return BAN_USER
    context.user_data['ban_user_id'] = int(user_id)
    await update.message.reply_text("Введите дату окончания блокировки (ДД.ММ.ГГГГ):")
    return BAN_DATE
#валидация даты
def validate_date(date_text):
    try:
        datetime.strptime(date_text, '%d.%m.%Y')
        return True
    except ValueError:
        return False
#ввод даты бана
async def ban_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = update.message.text.strip()
    if not validate_date(date_text):
        await update.message.reply_text(
            "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ:"
        )
        return BAN_DATE
    context.user_data['ban_end_date'] = date_text
    await update.message.reply_text("Укажите причину блокировки:")
    return BAN_REASON
#причина бана
async def ban_reason(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reason = update.message.text.strip()
    if not reason:
        await update.message.reply_text("Причина не может быть пустой. Укажите причину блокировки:")
        return BAN_REASON
    context.user_data['ban_reason'] = reason

    # Сохранение в Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_banned = wb["Banned"]
    ban_user_id = context.user_data['ban_user_id']

    # Получение ФИО пользователя из листа "Users"
    ws_users = wb["Users"]
    fio = None
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == ban_user_id:
            fio = row[1]
            break
    if not fio:
        await update.message.reply_text("Пользователь с таким ID не найден.")
        return ConversationHandler.END

    ws_banned.append([ban_user_id, fio, context.user_data['ban_end_date'], reason])
    wb.save(EXCEL_FILE)

    await update.message.reply_text("Пользователь успешно забанен!", reply_markup=main_menu(ban_user_id))
    return ConversationHandler.END
#отмена процесса бана
async def ban_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Процесс бана отменён.", reply_markup=main_menu(update.effective_user.id))
    return ConversationHandler.END
#открыт больничный
async def sick_leave_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие на кнопку 'Я открыл больничный'."""
    await update.message.reply_text("Введите дату следующего приёма (пример 21.10.2024):")
    return SICK_LEAVE_OPEN_DATE
#обработка даты приема
async def sick_leave_open_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод даты следующего приёма."""
    next_appointment_date = update.message.text.strip()

    user_id = update.effective_user.id
    # Получаем ФИО пользователя из листа "Users"
    fio = None

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            fio = row[1]
            break
    wb.close()

    if not fio:
        await update.message.reply_text("Не удалось найти ваше ФИО в базе данных.")
        return ConversationHandler.END

    # Формируем сообщение для отправки специальным пользователям
    message = f"{fio} открыл или продлил больничный, дата приема - {next_appointment_date}"

    # Отправляем сообщение специальным пользователям
    for special_user_id in SPECIAL_USER_IDS:
        try:
            await context.bot.send_message(chat_id=special_user_id, text=message)
        except Exception as e:
            logger.error(f"Не удалось отправить сообщение пользователю {special_user_id}: {e}")

    await update.message.reply_text("Информация отправлена.", reply_markup=main_menu(user_id))
    return ConversationHandler.END
#продлен больничный
async def sick_leave_return(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие на кнопку 'Больничный продлен'."""
    await update.message.reply_text("Введите дату нового приёма (пример 17.12.2024):")
    return SICK_LEAVE_OPEN_DATE
#закрыт больничный
async def sick_leave_close(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие на кнопку 'Я закрыл больничный'."""
    user_id = update.effective_user.id
    # Получаем ФИО пользователя из листа "Users"
    fio = None

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_users = wb["Users"]
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            fio = row[1]
            break
    wb.close()

    if not fio:
        await update.message.reply_text("Не удалось найти ваше ФИО в базе данных.")
        return

    # Формируем сообщение для отправки специальным пользователям
    message = f"{fio} закрыл больничный"

    # Отправляем сообщение специальным пользователям
    for special_user_id in SPECIAL_USER_IDS:
        try:
            await context.bot.send_message(chat_id=special_user_id, text=message)
        except Exception as e:
            logger.error(f"Не удалось отправить сообщение пользователю {special_user_id}: {e}")


    await update.message.reply_text("Информация отправлена ВС.", reply_markup=main_menu(user_id))
#скачать таблицу
async def download_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет файл data.xlsx специальному пользователю."""
    user_id = update.effective_user.id
    if user_id not in SPECIAL_USER_IDS and user_id not in get_older_users():
        await update.message.reply_text(
            "У вас нет доступа к этой функции.",
            reply_markup=main_menu(user_id)
        )
        return

    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text(
            "Файл не найден.",
            reply_markup=main_menu(user_id)
        )
        return

    with open(EXCEL_FILE, 'rb') as file:
        await update.message.reply_document(document=file, filename=EXCEL_FILE)

    await update.message.reply_text(
        "Файл отправлен.",
        reply_markup=main_menu(user_id)
    )
#очистка таблицы
async def clear_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатие на кнопку 'Очистить таблицу'."""
    user_id = update.effective_user.id

    if user_id not in SPECIAL_USER_IDS:
        await update.message.reply_text("У вас нет доступа к этой функции.")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            if sheet_name != 'Users':
                ws = wb[sheet_name]
                # Очищаем строки, начиная со второй, сохраняя заголовки
                ws.delete_rows(2, ws.max_row)

        wb.save(EXCEL_FILE)
        wb.close()

        await update.message.reply_text("Таблица успешно очищена, кроме данных на листе 'Users'.")
    except Exception as e:
        logger.error(f"Ошибка при очистке таблицы: {e}")
        await update.message.reply_text(f"Произошла ошибка при очистке таблицы: {e}")
#функция отмены
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена текущей операции."""
    await update.message.reply_text(
        'Операция отменена.',
        reply_markup=main_menu(update.effective_user.id)
    )
    return ConversationHandler.END
#старт расчета оплаты за час
async def calc_salary_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс расчета оплаты за час, устанавливая фиксированный оклад."""
    user_id = update.effective_user.id
    # Устанавливаем оклад в зависимости от того, является ли пользователь специальным
    if user_id in SPECIAL_USER_IDS:
        fixed_salary = 54125  # Фиксированный оклад для специальных пользователей
    else:
        fixed_salary = 40329.6  # Стандартный фиксированный оклад

    # Загружаем список OLDER_USERS
    older_users = get_older_users()
    if user_id in older_users:
        fixed_salary = 47845  # Фиксированный оклад для старших спецов

    context.user_data['salary'] = fixed_salary
    await update.message.reply_text(
        f"Фиксированный оклад установлен: {fixed_salary} P \n"
        "Введите норму часов в месяце: (узнать по ссылке https://www.consultant.ru/law/ref/calendar/proizvodstvennye/)",
        reply_markup=ReplyKeyboardRemove()
    )
    return CALC_HOURS
#расчет нормы часов и вычисление
async def calc_hours(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохраняет норму часов и выполняет расчет."""
    hours_text = update.message.text.strip()
    try:
        hours = float(hours_text.replace(',', '.'))
        if hours <= 0:
            raise ValueError
        context.user_data['hours'] = hours
    except ValueError:
        user_id = update.effective_user.id
        await update.message.reply_text(
            "Пожалуйста, введите правильное числовое значение нормы часов :"
        )
        return CALC_HOURS

    salary = context.user_data.get('salary')
    hours = context.user_data.get('hours')

    svu = ((salary * 0.87) / hours) * 1.5
    rvd = ((salary * 0.87) / hours) * 2
    svd = (rvd*8)
    svvd = (rvd * 10.75)

    # Форматирование чисел до 2 знаков после запятой
    svu = round(svu, 2)
    rvd = round(rvd, 2)
    svd = round(svd, 2)
    svvd = round(svvd, 2)

    message = (
        f"1 час СВУ стоит = {svu} ₽\n"
        f"1 час РВД стоит = {rvd} ₽\n\n"
        f"8 часовая смена РВД стоит = {svd} ₽\n"
        f"12 часовая смена РВД стоит = {svvd}₽\n"
        f"Не забывай о том, что стоит отдыхать"
    )

    user_id = update.effective_user.id
    await update.message.reply_text(message, reply_markup=main_menu(user_id))
    return ConversationHandler.END

#CАМОЕ ГЛАВНОЕ В РАБОТЕ БОТА
def main():
    """Основная функция для запуска бота."""
    # Получите ваш токен от @BotFather и вставьте ниже
    TOKEN = '7380924967:AAGbTShrh6-X59LY_sHX2NUCFvdOaNbzCwQ'  # Замените на ваш токен

    application = ApplicationBuilder().token(TOKEN).build()

    application.add_handler(MessageHandler(filters.Regex('^Проставленные активности$'), show_recorded_activities))
    # Обработчик Conversation для "Внести в бан"
    ban_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Внести в бан$"), ban_start)],
        states={
            BAN_USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, ban_user_id)],
            BAN_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ban_date)],
            BAN_REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, ban_reason)],
        },
        fallbacks=[CommandHandler('cancel', ban_cancel)],
    )

    # Обработчик Conversation для "Оповестить всех"
    notify_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Оповестить всех$"), notify_all_start)],
        states={
            NOTIFY_MESSAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, notify_all_message)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        allow_reentry=True
    )
    # ConversationHandler для регистрации
    register_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={

            REGISTER: [MessageHandler(filters.TEXT & ~filters.COMMAND, register_fio)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # ConversationHandler для удаления строки на листе "Activities"
    delete_any_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Удалить активити$'), delete_any_activity_start)],
        states={
            DELETE_ANY_ACTIVITY_ROW: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_any_activity_row)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # Conversation handler for transfer activity
    transfer_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Перенос активности$'), transfer_activity)],
        states={
            INTERVAL_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, interval_input)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        per_user=True,
        per_chat=True,
    )
    # ConversationHandler для добавления активности и закрашивания строк
    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Закрасить строки$"), color_rows_start)],
        states={
            COLOR_ROWS: [MessageHandler(filters.TEXT & ~filters.COMMAND, color_rows_process)],
        },
        fallbacks=[CommandHandler('cancel', lambda update, context: ConversationHandler.END)],
    )

    # ConversationHandler для добавления активности
    add_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^(Добавить активность)$'), add_activity_start)],
        states={
            DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_activity_date)],
            INTERVAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_activity_interval)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    # ConversationHandler для просмотра активностей
    view_activities_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^(Внесенные активности)$'), view_activities_start)],
        states={
            VIEW: [MessageHandler(filters.ALL, handle_message)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    # ConversationHandler для расчета оплаты за час
    calc_salary_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^(Рассчитать оплату за час)$'), calc_salary_start)],
        states={
            CALC_HOURS: [MessageHandler(filters.TEXT & ~filters.COMMAND, calc_hours)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # Обработчик для кнопки "Очистить таблицу"
    clear_table_handler = MessageHandler(
        filters.TEXT & filters.Regex('^Очистить таблицу$'),
        clear_table
    )
    # Обработчик для кнопки "Я открыл больничный"
    sick_leave_open_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Я открыл больничный$'), sick_leave_open)],
        states={
            SICK_LEAVE_OPEN_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, sick_leave_open_date)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    #Обработчик для кнопки "Больничный продлен"
    sick_leave_return_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Больничный продлен$'), sick_leave_open)],
        states={
            SICK_LEAVE_OPEN_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, sick_leave_open_date)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # ConversationHandler для редактирования активности
    edit_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Редактировать активность$"), edit_activity_start)],
        states={
            EDIT_SELECT_ACTIVITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_select_activity)],
            EDIT_FIELD: [MessageHandler(filters.Regex("^(Дата|Интервал)$"), edit_field)],
            EDIT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_value)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # ConversationHandler для удаления активности
    delete_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Удалить активность$"), delete_activity_start)],
        states={
            DELETE_SELECT_ACTIVITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_select_activity)],
            DELETE_CONFIRM: [MessageHandler(filters.Regex("^(Да|Нет)$"), delete_confirm)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    rename_handler_start = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^(Переименовать меня)$'), rename_handler)],
        states={
            RENAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, update_fio)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    training_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Отметить прохождение теста или обучения$"), activity_handler)],
        states={
            ACTIVITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, process_activity)],
            RESULT: [MessageHandler(filters.TEXT & ~filters.COMMAND, process_result)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    # ConversationHandler для ввода данных пользователя
    enter_data_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Внести данные$"), enter_data_start)],
        states={
            ENTER_DATA_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_data_id)],
            ENTER_DATA_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_data_fio)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # Обработчик для кнопки "Я закрыл больничный"
    sick_leave_close_handler = MessageHandler(
        filters.Regex('^Я закрыл больничный$'), sick_leave_close
    )
    delete_user_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Удалить пользователя$"), delete_user_start)],
        states={
            DELETE_USER_ROW: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_user_row)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    add_older_user_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Добавить старшего$"), add_older_user_start)],
        states={
            ADD_OLDER_USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_older_user_id)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    # ConversationHandler для удаления ID OLDER_USERS
    remove_older_user_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^Удалить старшего$"), remove_older_user_start)],
        states={
            REMOVE_OLDER_USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, remove_older_user_id)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    # ConversationHandler для ввода данных пользователя на лист "Activities"
    enter_activity_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^Внести активность$'), enter_activity_start)],
        states={
            ENTER_ACTIVITY_USER_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_activity_user_id)],
            ENTER_ACTIVITY_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_activity_fio)],
            ENTER_ACTIVITY_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_activity_date)],
            ENTER_ACTIVITY_INTERVAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_activity_interval)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )

    # Обработчик сообщений для главного меню

    application.add_handler(enter_activity_handler)
    application.add_handler(delete_any_activity_handler)
    application.add_handler(add_older_user_handler)
    application.add_handler(remove_older_user_handler)
    application.add_handler(delete_user_handler)
    application.add_handler(enter_data_handler)
    application.add_handler(training_activity_handler)
    application.add_handler(rename_handler_start)
    application.add_handler(ban_conv_handler)
    application.add_handler(notify_conv_handler)
    application.add_handler(sick_leave_return_handler)
    application.add_handler(sick_leave_open_handler)
    application.add_handler(conv_handler)
    application.add_handler(sick_leave_close_handler)
    application.add_handler(clear_table_handler)
    application.add_handler(transfer_activity_handler)
    application.add_handler(register_handler)
    application.add_handler(add_activity_handler)
    application.add_handler(view_activities_handler)
    application.add_handler(calc_salary_handler)
    application.add_handler(edit_activity_handler)
    application.add_handler(delete_activity_handler)
    application.add_handler(MessageHandler(filters.ATTACHMENT & ~filters.COMMAND, handle_document))
    application.job_queue.run_repeating(send_excel_file, interval=3600, first=10)

    # Добавляем обработчик для кнопок меню
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Добавляем обработчик для команды /cancel
    application.add_handler(CommandHandler('cancel', cancel))
    # Запуск бота
    application.run_polling()


if __name__ == '__main__':
    main()
